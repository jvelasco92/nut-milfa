"""
Generación de reportes PDF (ReportLab) y Excel (Pandas/OpenPyXL).
"""
from __future__ import annotations

import io

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image,
    PageBreak,
)

import branding
import utils_somatotype as som

PRIMARY_COLOR = colors.HexColor(branding.COLOR_PRIMARY)
PRIMARY_DARK = colors.HexColor(branding.COLOR_PRIMARY_DARK)
ACCENT_BRAND = colors.HexColor(branding.COLOR_ACCENT)
ACCENT_COLOR = colors.HexColor("#eaf3ee")
MUTED_COLOR = colors.HexColor(branding.COLOR_TEXT_MUTED)


def _dibujar_encabezado(canvas_obj, doc):
    """Header de marca (logo + nombre + nutricionistas) repetido en cada página."""
    canvas_obj.saveState()
    width, height = doc.pagesize
    y = height - 1.35 * cm

    r = 0.5 * cm
    cx = 1.9 * cm
    canvas_obj.setFillColor(PRIMARY_COLOR)
    canvas_obj.circle(cx, y, r, stroke=0, fill=1)

    canvas_obj.saveState()
    canvas_obj.translate(cx, y)
    canvas_obj.rotate(20)
    canvas_obj.setFillColor(ACCENT_BRAND)
    canvas_obj.ellipse(-r * 0.55, -r * 0.85, r * 0.55, r * 0.85, stroke=0, fill=1)
    canvas_obj.restoreState()

    canvas_obj.setFillColor(PRIMARY_DARK)
    canvas_obj.setFont("Helvetica-Bold", 15)
    canvas_obj.drawString(2.7 * cm, y + 0.12 * cm, branding.NOMBRE_APP)

    canvas_obj.setFillColor(ACCENT_BRAND)
    canvas_obj.setFont("Helvetica-Bold", 7)
    canvas_obj.drawString(2.7 * cm, y - 0.32 * cm, branding.TAGLINE)

    canvas_obj.setFillColor(MUTED_COLOR)
    canvas_obj.setFont("Helvetica", 8)
    canvas_obj.drawRightString(width - 1.5 * cm, y + 0.05 * cm, branding.NUTRICIONISTAS)

    canvas_obj.setStrokeColor(colors.HexColor("#e4e9ea"))
    canvas_obj.setLineWidth(0.8)
    canvas_obj.line(1.5 * cm, height - 2.05 * cm, width - 1.5 * cm, height - 2.05 * cm)

    canvas_obj.restoreState()


# ---------------------------------------------------------------------------
# Gráficos auxiliares (matplotlib) para embeber en el PDF
# ---------------------------------------------------------------------------
def _grafico_somatocarta_png(puntos: list[tuple[float, float]]) -> io.BytesIO:
    fig, ax = plt.subplots(figsize=(5, 4))
    ax.axhline(0, color="lightgray", linewidth=1)
    ax.axvline(0, color="lightgray", linewidth=1)
    ax.set_xlim(-9, 9)
    ax.set_ylim(-6, 12)
    ax.text(-8, -5.3, "ENDOMORFIA", fontsize=8, color="#c0392b")
    ax.text(6.2, -5.3, "ECTOMORFIA", fontsize=8, color="#2980b9")
    ax.text(-1.5, 11.2, "MESOMORFIA", fontsize=8, color="#27ae60")

    if puntos:
        xs = [p[0] for p in puntos]
        ys = [p[1] for p in puntos]
        if len(puntos) > 1:
            ax.plot(xs, ys, "--", color="gray", linewidth=1)
        ax.scatter(xs, ys, s=80, color=branding.COLOR_PRIMARY, edgecolor="black", zorder=5)
        for i, (x, y) in enumerate(zip(xs, ys), start=1):
            ax.annotate(str(i), (x, y), textcoords="offset points", xytext=(0, 6), ha="center", fontsize=8)

    ax.set_xlabel("X (Ecto - Endo)")
    ax.set_ylabel("Y (2·Meso - (Endo+Ecto))")
    ax.set_title("Somatocarta")
    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150)
    plt.close(fig)
    buf.seek(0)
    return buf


def _grafico_evolucion_png(df: pd.DataFrame) -> io.BytesIO:
    fig, ax = plt.subplots(figsize=(6, 3.2))
    ax.plot(df["fecha_medicion"], df["peso"], marker="o", label="Peso (kg)")
    ax2 = ax.twinx()
    ax2.plot(df["fecha_medicion"], df["porcentaje_grasa"], marker="s", color="#c0392b", label="% Grasa")
    ax2.plot(df["fecha_medicion"], df["porcentaje_musculo"], marker="^", color="#27ae60", label="% Músculo")

    ax.set_ylabel("Peso (kg)")
    ax2.set_ylabel("%")
    ax.set_xlabel("Fecha de medición")
    fig.autofmt_xdate(rotation=30)

    lines1, labels1 = ax.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax.legend(lines1 + lines2, labels1 + labels2, loc="upper left", fontsize=8)
    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150)
    plt.close(fig)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Estilos comunes
# ---------------------------------------------------------------------------
def _base_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TituloApp", fontSize=18, textColor=PRIMARY_COLOR, spaceAfter=4, leading=22))
    styles.add(ParagraphStyle(name="Subtitulo", fontSize=11, textColor=colors.grey, spaceAfter=12))
    styles.add(ParagraphStyle(name="Seccion", fontSize=13, textColor=PRIMARY_COLOR, spaceBefore=14, spaceAfter=6))
    return styles


def _tabla_datos(pairs: list[tuple[str, str]], col_widths=(6 * cm, 6 * cm)) -> Table:
    data = [[k, v] for k, v in pairs]
    t = Table(data, colWidths=list(col_widths))
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), ACCENT_COLOR),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return t


def _fmt(v, suf=""):
    if v is None or v == "" or (isinstance(v, float) and pd.isna(v)):
        return "-"
    return f"{v}{suf}"


def _tabla_semaforo(filas: list[tuple[str, str, str, str]], col_widths=(6.5 * cm, 3 * cm, 3.5 * cm)) -> Table:
    """filas: (nombre_metrica, valor_texto, etiqueta_clasificación, color_hex)."""
    data = [["Métrica", "Valor", "Referencia"]] + [[n, v, et] for n, v, et, _ in filas]
    t = Table(data, colWidths=list(col_widths))
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), PRIMARY_COLOR),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]
    for i, (_, _, _, color_hex) in enumerate(filas, start=1):
        bg = colors.HexColor(color_hex)
        texto = colors.black if color_hex.upper() == som.COLOR_AMARILLO.upper() else colors.white
        style += [
            ("BACKGROUND", (2, i), (2, i), bg),
            ("TEXTCOLOR", (2, i), (2, i), texto),
            ("FONTNAME", (2, i), (2, i), "Helvetica-Bold"),
        ]
    t.setStyle(TableStyle(style))
    return t


# ---------------------------------------------------------------------------
# PDF: Ficha individual
# ---------------------------------------------------------------------------
def _delta_texto(actual, anterior, suf=""):
    if actual is None or anterior is None or pd.isna(actual) or pd.isna(anterior):
        return "-"
    diff = round(actual - anterior, 2)
    signo = "+" if diff >= 0 else ""
    return f"{signo}{diff}{suf}"


def generar_pdf_ficha_individual(atleta: dict, medicion: dict, medicion_anterior: dict = None) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2.6 * cm, bottomMargin=1.5 * cm)
    styles = _base_styles()
    story = []
    sexo = atleta.get("sexo", "Masculino")

    story.append(Paragraph("Ficha Antropométrica Individual", styles["TituloApp"]))
    story.append(
        Paragraph(
            f"{atleta['nombre']} {atleta['apellido']} — Grupo: {atleta.get('nombre_grupo', '-')}",
            styles["Subtitulo"],
        )
    )

    edad_al_momento = atleta.get("edad")
    if atleta.get("fecha_nacimiento") and medicion.get("fecha_medicion"):
        edad_al_momento = som.calcular_edad(atleta["fecha_nacimiento"], medicion["fecha_medicion"])

    story.append(_tabla_datos([
        ("Sexo", _fmt(sexo)),
        ("Fecha de nacimiento", _fmt(atleta.get("fecha_nacimiento"))),
        ("Edad al momento de la medición", _fmt(edad_al_momento, " años")),
        ("Email", _fmt(atleta.get("email"))),
        ("Fecha de medición", _fmt(medicion.get("fecha_medicion"))),
        ("Fecha y hora de carga", _fmt(medicion.get("fecha_hora_carga"))),
        ("Cargado por", _fmt(medicion.get("cargado_por"))),
    ]))

    # --- Datos principales con semáforo de referencia (siempre se muestran) ---
    story.append(Paragraph("Datos Principales", styles["Seccion"]))
    metricas_core = [
        ("Peso (kg)", medicion.get("peso"), None, "peso"),
        ("Talla (m)", (medicion.get("altura") / 100) if medicion.get("altura") else None, None, "talla"),
        ("IMC (kg/m²)", medicion.get("imc"), "imc", "imc"),
        ("% Músculo esquelético estimado", medicion.get("pct_musculo_esqueletico"), "musculo_esqueletico_pct", "musculo"),
        ("% Grasa corporal estimado", medicion.get("bio_grasa_corporal"), "grasa_corporal_pct", "grasa"),
        ("% Grasa visceral estimada", medicion.get("bio_grasa_visceral"), "grasa_visceral", "visceral"),
        ("Circ. cintura (cm)", medicion.get("cintura"), "circ_cintura", "cm"),
        ("Circ. cadera (cm)", medicion.get("cadera"), None, "cm"),
        ("Pliegue abdominal (mm)", medicion.get("pliegue_abdominal"), "pliegue_abdominal", "mm"),
        ("Índice cintura/cadera", medicion.get("indice_cintura_cadera"), "indice_cintura_cadera", ""),
        ("Índice cintura/talla", medicion.get("indice_cintura_talla"), "indice_cintura_talla", ""),
    ]
    filas = []
    for nombre, valor, metrica_ref, _ in metricas_core:
        if metrica_ref:
            etiqueta, color = som.clasificar_metrica(metrica_ref, valor, sexo)
        else:
            etiqueta, color = ("—", branding.COLOR_TEXT_MUTED)
        filas.append((nombre, _fmt(round(valor, 2) if isinstance(valor, float) else valor), etiqueta, color))
    story.append(_tabla_semaforo(filas))

    if sexo == "Femenino":
        story.append(Spacer(1, 4))
        story.append(Paragraph(
            "Circ. cintura, índice cintura/cadera, índice cintura/talla y % músculo esquelético no tienen "
            "referencia cargada para mujeres todavía; se muestran sin clasificar.",
            styles["Normal"],
        ))

    # --- Comparación con la medición anterior ---
    if medicion_anterior:
        story.append(Paragraph("Comparación con Medición Anterior", styles["Seccion"]))
        story.append(Paragraph(
            f"Anterior: {_fmt(medicion_anterior.get('fecha_medicion'))}  →  "
            f"Actual: {_fmt(medicion.get('fecha_medicion'))}",
            styles["Normal"],
        ))
        story.append(Spacer(1, 4))
        story.append(_tabla_datos([
            ("Δ Peso (kg)", _delta_texto(medicion.get("peso"), medicion_anterior.get("peso"))),
            ("Δ IMC", _delta_texto(medicion.get("imc"), medicion_anterior.get("imc"))),
            ("Δ % Músculo esquelético", _delta_texto(medicion.get("pct_musculo_esqueletico"), medicion_anterior.get("pct_musculo_esqueletico"), " %")),
            ("Δ % Grasa corporal", _delta_texto(medicion.get("bio_grasa_corporal"), medicion_anterior.get("bio_grasa_corporal"), " %")),
            ("Δ Circ. cintura", _delta_texto(medicion.get("cintura"), medicion_anterior.get("cintura"), " cm")),
        ]))

    # --- Secciones ISAK avanzadas: solo si están cargadas ---
    hay_perimetros = any([
        medicion.get("brazo_relajado"), medicion.get("brazo_contraido"),
        medicion.get("perimetro_muslo_medio"), medicion.get("perimetro_pantorrilla"),
    ])
    if hay_perimetros:
        story.append(Paragraph("Perímetros Adicionales (cm)", styles["Seccion"]))
        story.append(_tabla_datos([
            ("Brazo relajado", _fmt(medicion.get("brazo_relajado"))),
            ("Brazo contraído", _fmt(medicion.get("brazo_contraido"))),
            ("Muslo medio", _fmt(medicion.get("perimetro_muslo_medio"))),
            ("Pantorrilla", _fmt(medicion.get("perimetro_pantorrilla"))),
        ]))

    hay_pliegues_isak = any([
        medicion.get("pliegue_tricipital"), medicion.get("pliegue_subescapular"),
        medicion.get("pliegue_suprailiaco"), medicion.get("pliegue_muslo_medio"),
        medicion.get("pliegue_pantorrilla"),
    ])
    if hay_pliegues_isak:
        story.append(Paragraph("Pliegues Cutáneos ISAK (mm)", styles["Seccion"]))
        story.append(_tabla_datos([
            ("Tricipital", _fmt(medicion.get("pliegue_tricipital"))),
            ("Subescapular", _fmt(medicion.get("pliegue_subescapular"))),
            ("Suprailíaco", _fmt(medicion.get("pliegue_suprailiaco"))),
            ("Abdominal", _fmt(medicion.get("pliegue_abdominal"))),
            ("Muslo medio", _fmt(medicion.get("pliegue_muslo_medio"))),
            ("Pantorrilla", _fmt(medicion.get("pliegue_pantorrilla"))),
            ("Sumatoria 6 pliegues", _fmt(medicion.get("sumatoria_6_pliegues"), " mm")),
            ("% Grasa (Yuhasz, por pliegues)", _fmt(medicion.get("porcentaje_grasa"), " %")),
            ("% Músculo (Martin, por perímetros)", _fmt(medicion.get("porcentaje_musculo"), " %")),
        ]))

    hay_somatotipo = any([medicion.get("endomorfia"), medicion.get("mesomorfia"), medicion.get("ectomorfia")])
    if hay_somatotipo:
        story.append(Paragraph("Somatotipo (Heath-Carter)", styles["Seccion"]))
        story.append(_tabla_datos([
            ("Endomorfia", _fmt(medicion.get("endomorfia"))),
            ("Mesomorfia", _fmt(medicion.get("mesomorfia"))),
            ("Ectomorfia", _fmt(medicion.get("ectomorfia"))),
            ("Coordenadas (X, Y)", f"({_fmt(medicion.get('coord_x'))}, {_fmt(medicion.get('coord_y'))})"),
        ]))
        x, y = medicion.get("coord_x"), medicion.get("coord_y")
        if x is not None and y is not None:
            story.append(Spacer(1, 10))
            story.append(Image(_grafico_somatocarta_png([(x, y)]), width=11 * cm, height=8.8 * cm))

    hay_bio_adicional = any([
        medicion.get("bio_agua_corporal"), medicion.get("bio_masa_muscular"),
        medicion.get("bio_masa_osea"), medicion.get("bio_metabolismo_basal"), medicion.get("bio_edad_metabolica"),
    ])
    if hay_bio_adicional:
        story.append(Paragraph("Bioimpedancia Adicional", styles["Seccion"]))
        story.append(_tabla_datos([
            ("% Agua corporal", _fmt(medicion.get("bio_agua_corporal"), " %")),
            ("Masa muscular", _fmt(medicion.get("bio_masa_muscular"), " kg")),
            ("Masa ósea", _fmt(medicion.get("bio_masa_osea"), " kg")),
            ("Metabolismo basal", _fmt(medicion.get("bio_metabolismo_basal"), " kcal")),
            ("Edad metabólica", _fmt(medicion.get("bio_edad_metabolica"), " años")),
        ]))

    story.append(Paragraph("Observación", styles["Seccion"]))
    obs = medicion.get("observaciones") or "Sin observaciones registradas."
    story.append(Paragraph(obs, styles["Normal"]))

    doc.build(story, onFirstPage=_dibujar_encabezado, onLaterPages=_dibujar_encabezado)
    buffer.seek(0)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# PDF: Historial
# ---------------------------------------------------------------------------
def generar_pdf_historial(atleta: dict, df_historial: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2.6 * cm, bottomMargin=1.5 * cm)
    styles = _base_styles()
    story = []

    story.append(Paragraph("Historial Antropométrico", styles["TituloApp"]))
    story.append(
        Paragraph(
            f"{atleta['nombre']} {atleta['apellido']} — Grupo: {atleta.get('nombre_grupo', '-')}",
            styles["Subtitulo"],
        )
    )

    cols = ["fecha_medicion", "peso", "imc", "porcentaje_grasa", "porcentaje_musculo", "sumatoria_6_pliegues"]
    headers = ["Fecha", "Peso", "IMC", "% Grasa", "% Músculo", "Σ6 Pliegues"]
    data = [headers] + df_historial[cols].astype(str).values.tolist()
    t = Table(data, colWidths=[3 * cm] + [2.6 * cm] * 5)
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), PRIMARY_COLOR),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ACCENT_COLOR]),
                ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ]
        )
    )
    story.append(t)

    if len(df_historial) >= 2:
        story.append(Paragraph("Evolución Temporal", styles["Seccion"]))
        story.append(Image(_grafico_evolucion_png(df_historial), width=15 * cm, height=8 * cm))

    puntos = list(zip(df_historial["coord_x"].dropna(), df_historial["coord_y"].dropna()))
    if puntos:
        story.append(Paragraph("Somatocarta - Evolución", styles["Seccion"]))
        story.append(Image(_grafico_somatocarta_png(puntos), width=11 * cm, height=8.8 * cm))

    doc.build(story, onFirstPage=_dibujar_encabezado, onLaterPages=_dibujar_encabezado)
    buffer.seek(0)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def generar_excel_historial(atleta: dict, df_historial: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_historial.to_excel(writer, index=False, sheet_name="Historial")
        info = pd.DataFrame(
            [{"Nombre": atleta["nombre"], "Apellido": atleta["apellido"],
              "Sexo": atleta.get("sexo"), "Edad": atleta.get("edad"),
              "Grupo": atleta.get("nombre_grupo")}]
        )
        info.to_excel(writer, index=False, sheet_name="Atleta")
    buffer.seek(0)
    return buffer.getvalue()


_COLUMNAS_GRUPO = [
    ("nombre", "Nombre", None),
    ("apellido", "Apellido", None),
    ("fecha_medicion", "Fecha medición", None),
    ("peso", "Peso (kg)", None),
    ("talla_m", "Talla (m)", None),
    ("imc", "IMC (kg/m²)", "imc"),
    ("pct_musculo_esqueletico", "% Músculo esquelético", "musculo_esqueletico_pct"),
    ("bio_grasa_corporal", "% Grasa corporal", "grasa_corporal_pct"),
    ("bio_grasa_visceral", "% Grasa visceral", "grasa_visceral"),
    ("cintura", "Circ. cintura (cm)", "circ_cintura"),
    ("cadera", "Circ. cadera (cm)", None),
    ("pliegue_abdominal", "Pliegue abdominal (mm)", "pliegue_abdominal"),
    ("indice_cintura_cadera", "Índice cintura/cadera", "indice_cintura_cadera"),
    ("indice_cintura_talla", "Índice cintura/talla", "indice_cintura_talla"),
]


def generar_excel_grupal(nombre_grupo: str, df_detalle: pd.DataFrame, estadisticas: dict) -> bytes:
    """Tabla coloreada (semáforo) por atleta, con leyenda de referencias y promedios,
    inspirada en la planilla de las nutricionistas."""
    df = df_detalle.copy()
    df["talla_m"] = df["altura"].apply(lambda v: round(v / 100, 2) if pd.notna(v) else None)

    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle"

    fill_titulo = PatternFill("solid", fgColor="1F6F78")
    fill_header = PatternFill("solid", fgColor="FFE599")
    font_titulo = Font(bold=True, color="FFFFFF", size=13)
    font_header = Font(bold=True, size=10)

    n_cols = len(_COLUMNAS_GRUPO)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(row=1, column=1, value=f"CONTROL DE COMPOSICIÓN CORPORAL — {nombre_grupo}").font = font_titulo
    ws.cell(row=1, column=1).fill = fill_titulo
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.cell(row=2, column=1, value=branding.NUTRICIONISTAS).font = Font(italic=True, size=10)

    header_row = 4
    for col_idx, (_, titulo, _) in enumerate(_COLUMNAS_GRUPO, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=titulo)
        c.font = font_header
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, (_, fila) in enumerate(df.iterrows(), start=header_row + 1):
        sexo = fila.get("sexo", "Masculino")
        for col_idx, (campo, _, metrica_ref) in enumerate(_COLUMNAS_GRUPO, start=1):
            valor = fila.get(campo)
            valor = None if pd.isna(valor) else valor
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.alignment = Alignment(horizontal="center")
            if metrica_ref:
                _, color = som.clasificar_metrica(metrica_ref, valor, sexo)
                if color not in (som.COLOR_SIN_DATO,):
                    cell.fill = PatternFill("solid", fgColor=color.lstrip("#").upper())

    for col_idx in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    # --- Hoja de referencias (leyenda de colores y umbrales) ---
    ws_ref = wb.create_sheet("Referencias")
    ws_ref.cell(row=1, column=1, value="Referencias utilizadas para el semáforo de colores").font = Font(bold=True, size=12)
    fila_actual = 3
    etiquetas_metricas = {
        "imc": "IMC (kg/m²)",
        "grasa_corporal_pct": "% Grasa corporal",
        "grasa_visceral": "% Grasa visceral",
        "pliegue_abdominal": "Pliegue abdominal (mm)",
        "circ_cintura": "Circ. cintura (cm) — solo hombres",
        "indice_cintura_cadera": "Índice cintura/cadera — solo hombres",
        "indice_cintura_talla": "Índice cintura/talla — solo hombres",
        "musculo_esqueletico_pct": "% Músculo esquelético (18-39 años) — solo hombres",
    }
    for metrica, titulo in etiquetas_metricas.items():
        ws_ref.cell(row=fila_actual, column=1, value=titulo).font = Font(bold=True)
        fila_actual += 1
        reglas = som.reglas_referencia(metrica)
        for minimo, maximo, etiqueta, color in reglas:
            if minimo is None:
                rango = f"< {maximo}"
            elif maximo is None:
                rango = f"≥ {minimo}"
            else:
                rango = f"{minimo} - {maximo}"
            c = ws_ref.cell(row=fila_actual, column=1, value=f"{etiqueta}: {rango}")
            c.fill = PatternFill("solid", fgColor=color.lstrip("#").upper())
            fila_actual += 1
        fila_actual += 1
    ws_ref.column_dimensions["A"].width = 45

    # --- Hoja de promedios grupales ---
    ws_prom = wb.create_sheet("Promedios")
    etiquetas_prom = {
        "cantidad_atletas": "Atletas con medición",
        "peso_prom": "Peso promedio (kg)",
        "imc_prom": "IMC promedio",
        "grasa_prom": "% Grasa corporal promedio",
        "musculo_prom": "% Músculo esquelético promedio",
        "grasa_visceral_prom": "% Grasa visceral promedio",
        "icc_prom": "Índice cintura/cadera promedio",
        "ict_prom": "Índice cintura/talla promedio",
        "sumatoria_prom": "Sumatoria 6 pliegues promedio (ISAK)",
        "endomorfia_prom": "Endomorfia grupal (ISAK)",
        "mesomorfia_prom": "Mesomorfia grupal (ISAK)",
        "ectomorfia_prom": "Ectomorfia grupal (ISAK)",
    }
    ws_prom.cell(row=1, column=1, value="Métrica").font = Font(bold=True)
    ws_prom.cell(row=1, column=2, value="Promedio").font = Font(bold=True)
    for i, (clave, etiqueta) in enumerate(etiquetas_prom.items(), start=2):
        valor = estadisticas.get(clave)
        valor = round(valor, 2) if isinstance(valor, float) and pd.notna(valor) else valor
        ws_prom.cell(row=i, column=1, value=etiqueta)
        ws_prom.cell(row=i, column=2, value=valor)
    ws_prom.column_dimensions["A"].width = 38
    ws_prom.column_dimensions["B"].width = 14

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# PDF: Estadística de grupo
# ---------------------------------------------------------------------------
_ANCHOS_COLUMNAS_GRUPO_PDF = [2.3, 2.3, 2.1, 1.5, 1.3, 1.5, 1.9, 1.7, 1.7, 1.8, 1.6, 1.9, 1.7, 1.7]  # cm


def generar_pdf_grupal(nombre_grupo: str, df_detalle: pd.DataFrame, estadisticas: dict) -> bytes:
    """Reporte PDF de estadística grupal: promedios + tabla por atleta con semáforo."""
    df = df_detalle.copy()
    df["talla_m"] = df["altura"].apply(lambda v: round(v / 100, 2) if pd.notna(v) else None)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=2.6 * cm, bottomMargin=1.5 * cm,
                             leftMargin=1.3 * cm, rightMargin=1.3 * cm)
    styles = _base_styles()
    story = []

    story.append(Paragraph(f"Estadística de Grupo: {nombre_grupo}", styles["TituloApp"]))

    def _r(v):
        return round(v, 2) if isinstance(v, (int, float)) and pd.notna(v) else "-"

    story.append(Paragraph("Promedios del Grupo", styles["Seccion"]))
    story.append(_tabla_datos([
        ("Atletas con medición", _fmt(int(estadisticas.get("cantidad_atletas") or 0))),
        ("IMC promedio", _fmt(_r(estadisticas.get("imc_prom")))),
        ("% Grasa corporal promedio", _fmt(_r(estadisticas.get("grasa_prom")), " %" if estadisticas.get("grasa_prom") else "")),
        ("% Músculo esquelético promedio", _fmt(_r(estadisticas.get("musculo_prom")), " %" if estadisticas.get("musculo_prom") else "")),
        ("% Grasa visceral promedio", _fmt(_r(estadisticas.get("grasa_visceral_prom")))),
        ("Índice cintura/cadera promedio", _fmt(_r(estadisticas.get("icc_prom")))),
        ("Índice cintura/talla promedio", _fmt(_r(estadisticas.get("ict_prom")))),
    ], col_widths=(8 * cm, 5 * cm)))

    story.append(Paragraph("Detalle por Atleta (última medición)", styles["Seccion"]))
    headers = [titulo for _, titulo, _ in _COLUMNAS_GRUPO]
    data = [headers]
    filas_color = []
    for _, fila in df.iterrows():
        sexo = fila.get("sexo", "Masculino")
        fila_valores = []
        fila_colores = []
        for campo, _, metrica_ref in _COLUMNAS_GRUPO:
            valor = fila.get(campo)
            valor = None if pd.isna(valor) else valor
            fila_valores.append(_fmt(valor))
            if metrica_ref:
                _, color = som.clasificar_metrica(metrica_ref, valor, sexo)
                fila_colores.append(None if color == som.COLOR_SIN_DATO else color)
            else:
                fila_colores.append(None)
        data.append(fila_valores)
        filas_color.append(fila_colores)

    col_widths = [w * cm for w in _ANCHOS_COLUMNAS_GRUPO_PDF]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), PRIMARY_COLOR),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
        ("ALIGN", (2, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for r_idx, fila_colores in enumerate(filas_color, start=1):
        for c_idx, color_hex in enumerate(fila_colores):
            if color_hex:
                bg = colors.HexColor(color_hex)
                texto = colors.black if color_hex.upper() == som.COLOR_AMARILLO.upper() else colors.white
                style += [
                    ("BACKGROUND", (c_idx, r_idx), (c_idx, r_idx), bg),
                    ("TEXTCOLOR", (c_idx, r_idx), (c_idx, r_idx), texto),
                    ("FONTNAME", (c_idx, r_idx), (c_idx, r_idx), "Helvetica-Bold"),
                ]
    t.setStyle(TableStyle(style))
    story.append(t)

    doc.build(story, onFirstPage=_dibujar_encabezado, onLaterPages=_dibujar_encabezado)
    buffer.seek(0)
    return buffer.getvalue()
